import streamlit as st
import openpyxl
from io import BytesIO
import tempfile

st.title("💹 改变表格中数值单位")

with st.expander("💡 使用介绍"):
    st.info("首先上传 xlsx 或者 csv 文件，然后选择文件中需要调整单位的表格，输入更改后的单位。例如：万元 输入 10000。最后点击下载。")
    st.info("注意：这里默认原上传文件单位为元，如果原上传单位为万元，需要调整到元，则在‘更改后单位值’输入 0.0001。")

uploaded_file = st.file_uploader("1. 上传表格文件", type=["csv","xlsx"])

if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file)

    container = st.container()
    all = st.checkbox("选择所有")

    if all:
        selected_options = container.multiselect("2. 选择一个或者多个表格:",
            wb.sheetnames,wb.sheetnames)
    else:
        selected_options =  container.multiselect("2. 选择一个或者多个表格:",
            wb.sheetnames)

    d = st.text_input("3. 更改后单位值")

    if len(selected_options) != 0 and d != '':
        for i in selected_options:
            ws = wb[i]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'n':
                        if cell.value != None:
                            cell.value = cell.value/int(d)
        
        with tempfile.NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            data = BytesIO(tmp.read())

        st.download_button(label='📥 下载结果文件', data=data, mime='xlsx', file_name= uploaded_file.name)